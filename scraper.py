"""
Web Scraper Demo — books.toscrape.com
Demonstrates: pagination, data extraction, multi-format export.
"""

import requests
from bs4 import BeautifulSoup
import csv
import json
import argparse
import time
import os


HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                   "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
}

RATING_MAP = {"One": 1, "Two": 2, "Three": 3, "Four": 4, "Five": 5}


def fetch_page(url, delay=1.0):
    """Fetch and parse a page with polite delay."""
    time.sleep(delay)
    resp = requests.get(url, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    return BeautifulSoup(resp.text, "html.parser")


def extract_books(soup):
    """Extract book data from a catalogue page."""
    books = []
    for article in soup.select("article.product_pod"):
        title = article.select_one("h3 a")["title"]
        price_text = article.select_one(".price_color").get_text(strip=True)
        price = float(price_text.replace("£", "").replace("Â", ""))
        rating_class = article.select_one("p.star-rating")["class"][1]
        rating = RATING_MAP.get(rating_class, 0)
        in_stock = "In stock" in article.select_one(".availability").get_text()
        detail_url = article.select_one("h3 a")["href"]

        books.append({
            "title": title,
            "price_gbp": price,
            "rating": rating,
            "in_stock": in_stock,
            "detail_url": detail_url,
        })
    return books


def scrape_catalogue(base_url, max_pages=5):
    """Scrape multiple catalogue pages."""
    all_books = []
    for page in range(1, max_pages + 1):
        url = f"{base_url}/catalogue/page-{page}.html"
        print(f"[{page}/{max_pages}] {url}")
        try:
            soup = fetch_page(url)
            books = extract_books(soup)
            if not books:
                print(f"  No books found, stopping.")
                break
            all_books.extend(books)
            print(f"  Found {len(books)} books (total: {len(all_books)})")
        except requests.HTTPError as e:
            print(f"  HTTP error: {e}")
            break
    return all_books


def save_csv(data, filepath):
    if not data:
        return
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=data[0].keys())
        w.writeheader()
        w.writerows(data)
    print(f"Saved {len(data)} rows → {filepath}")


def save_json(data, filepath):
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"Saved {len(data)} items → {filepath}")


def save_excel(data, filepath):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("openpyxl not installed — falling back to CSV")
        save_csv(data, filepath.replace(".xlsx", ".csv"))
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Books"

    headers = list(data[0].keys())
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E86AB")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, record in enumerate(data, 2):
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=record[h])

    for col_idx, h in enumerate(headers, 1):
        max_len = max(len(str(h)), *(len(str(r[h])) for r in data[:50]))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 3, 50)

    ws.auto_filter.ref = ws.dimensions
    wb.save(filepath)
    print(f"Saved {len(data)} rows → {filepath}")


def main():
    parser = argparse.ArgumentParser(description="Book Scraper Demo")
    parser.add_argument("--pages", type=int, default=3, help="Pages to scrape (default 3)")
    parser.add_argument("--output", default="sample_output/books_sample.csv",
                        help="Output file (.csv, .json, or .xlsx)")
    args = parser.parse_args()

    base = "https://books.toscrape.com"
    books = scrape_catalogue(base, args.pages)

    os.makedirs(os.path.dirname(args.output) or ".", exist_ok=True)

    ext = os.path.splitext(args.output)[1].lower()
    if ext == ".json":
        save_json(books, args.output)
    elif ext == ".xlsx":
        save_excel(books, args.output)
    else:
        save_csv(books, args.output)

    print(f"\nDone! Scraped {len(books)} books.")


if __name__ == "__main__":
    main()
